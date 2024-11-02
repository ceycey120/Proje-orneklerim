import tkinter as tk
from tkinter import messagebox
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from tkinter import filedialog
import logging

# Loglama ayarları
logging.basicConfig(filename='error.log', level=logging.ERROR)

gelirler = []
giderler = []

# Gelir ekleme fonksiyonu
def gelir_ekle():
    try:
        gelir_turu = gelir_turu_entry.get()
        gelir_miktari = float(gelir_miktari_entry.get())
        gelirler.append({"Tür": gelir_turu, "Miktar": gelir_miktari})
        messagebox.showinfo("Başarılı", "Gelir başarıyla eklendi.")
    except ValueError:
        messagebox.showerror("Hata", "Geçerli bir miktar girin.")
    gelir_turu_entry.delete(0, tk.END)
    gelir_miktari_entry.delete(0, tk.END)

# Gider ekleme fonksiyonu
def gider_ekle():
    try:
        gider_turu = gider_turu_entry.get()
        gider_miktari = float(gider_miktari_entry.get())
        giderler.append({"Tür": gider_turu, "Miktar": gider_miktari})
        messagebox.showinfo("Başarılı", "Gider başarıyla eklendi.")
    except ValueError:
        messagebox.showerror("Hata", "Geçerli bir miktar girin.")
    gider_turu_entry.delete(0, tk.END)
    gider_miktari_entry.delete(0, tk.END)

# Hesaplama fonksiyonu
def hesapla():
    toplam_gelir = sum([gelir["Miktar"] for gelir in gelirler])
    toplam_gider = sum([gider["Miktar"] for gider in giderler])
    net_gelir = toplam_gelir - toplam_gider

    sonuc_label.config(text=f"Toplam Gelir: {toplam_gelir} TL\nToplam Gider: {toplam_gider} TL\nNet Gelir: {net_gelir} TL")

# Excel dosyasına kaydetme fonksiyonu
def excel_kaydet():
    try:
        if not gelirler and not giderler:
            messagebox.showerror("Hata", "Kaydedilecek veri yok.")
            return

        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not filepath:
            return

        # Gelir ve Gider DataFrames
        gelirler_df = pd.DataFrame(gelirler)
        giderler_df = pd.DataFrame(giderler)

        # Yeni Excel dosyası oluşturma
        wb = Workbook()
        ws = wb.active
        ws.title = "Gelir-Gider Raporu"

        # Başlıklar
        headers = ["Gelir Türü", "Gelir Miktarı", "Gider Türü", "Gider Miktarı"]
        ws.append(headers)
        header_fill = PatternFill(start_color="D9B9FF", end_color="D9B9FF", fill_type="solid")
        header_font = Font(bold=True, color="000000")
        header_alignment = Alignment(horizontal="center")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Verileri ekleme
        max_len = max(len(gelirler), len(giderler))
        for i in range(max_len):
            gelir_turu = gelirler[i]["Tür"] if i < len(gelirler) else ""
            gelir_miktari = gelirler[i]["Miktar"] if i < len(gelirler) else ""
            gider_turu = giderler[i]["Tür"] if i < len(giderler) else ""
            gider_miktari = giderler[i]["Miktar"] if i < len(giderler) else ""

            row_data = [gelir_turu, gelir_miktari, gider_turu, gider_miktari]
            ws.append(row_data)

            # Hücreler ortalanmış şekilde ayarlandı
            for j in range(1, 5):
                ws.cell(row=i + 2, column=j).alignment = Alignment(horizontal="center")

        # Toplam Satırı
        toplam_gelir = sum([gelir["Miktar"] for gelir in gelirler])
        toplam_gider = sum([gider["Miktar"] for gider in giderler])
        net_gelir = toplam_gelir - toplam_gider

        ws.append(["Toplam", toplam_gelir, "Toplam", toplam_gider])
        ws[f"A{max_len + 2}"].font = Font(bold=True, color="000000")
        ws[f"A{max_len + 2}"].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        ws[f"B{max_len + 2}"].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        ws[f"C{max_len + 2}"].font = Font(bold=True, color="000000")
        ws[f"C{max_len + 2}"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        ws[f"D{max_len + 2}"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # Net Gelir Satırı
        ws.merge_cells(f"A{max_len + 3}:C{max_len + 3}")
        ws[f"A{max_len + 3}"] = "Net Gelir"
        ws[f"A{max_len + 3}"].font = Font(bold=True)
        ws[f"A{max_len + 3}"].alignment = Alignment(horizontal="center")
        ws[f"A{max_len + 3}"].fill = PatternFill(start_color="00CCFF", end_color="00CCFF", fill_type="solid")

        ws[f"D{max_len + 3}"] = net_gelir
        ws[f"D{max_len + 3}"].font = Font(bold=True)
        ws[f"D{max_len + 3}"].alignment = Alignment(horizontal="center")
        ws[f"D{max_len + 3}"].fill = PatternFill(start_color="00CCFF", end_color="00CCFF", fill_type="solid")

        # Hücre çerçevesi ekleme
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border

        # Otomatik sütun genişliği
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col_letter].width = adjusted_width

        wb.save(filepath)
        messagebox.showinfo("Başarılı", "Veriler Excel dosyasına başarıyla kaydedildi.")
    except Exception as e:
        logging.error(str(e))
        messagebox.showerror("Hata", f"Excel kaydedilirken bir hata oluştu: {e}")

# Arayüz 
root = tk.Tk()
root.title("Gelir ve Gider Hesaplama Aracı")

# Gelir Ekleme 
tk.Label(root, text="Gelir Türü:").grid(row=0, column=0)
gelir_turu_entry = tk.Entry(root)
gelir_turu_entry.grid(row=0, column=1)

tk.Label(root, text="Gelir Miktarı:").grid(row=1, column=0)
gelir_miktari_entry = tk.Entry(root)
gelir_miktari_entry.grid(row=1, column=1)

tk.Button(root, text="Gelir Ekle", command=gelir_ekle).grid(row=2, column=1)

# Gider Ekleme 
tk.Label(root, text="Gider Türü:").grid(row=3, column=0)
gider_turu_entry = tk.Entry(root)
gider_turu_entry.grid(row=3, column=1)

tk.Label(root, text="Gider Miktarı:").grid(row=4, column=0)
gider_miktari_entry = tk.Entry(root)
gider_miktari_entry.grid(row=4, column=1)

tk.Button(root, text="Gider Ekle", command=gider_ekle).grid(row=5, column=1)

# Hesaplama 
tk.Button(root, text="Hesapla", command=hesapla).grid(row=6, column=1)
sonuc_label = tk.Label(root, text="")
sonuc_label.grid(row=7, column=1)

# Excel'e Kaydet
tk.Button(root, text="Excel'e Kaydet", command=excel_kaydet).grid(row=8, column=1)

root.mainloop()
